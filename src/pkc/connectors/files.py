"""Datei-Connectoren: CSV und Excel.

Diese beiden sind vollstaendig implementiert und sofort nutzbar - sie
brauchen kein fremdes System, nur eine Datei auf der SSD.  Damit ist der
haeufigste Praxisfall (Export aus dem Buchhaltungssystem) abgedeckt.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from .base import Connector, ConnectorError, ConnectorMode, ConnectorResult, NotConfigured


def resolve_within(directory: Path, name: str) -> Path:
    """Loest einen Dateinamen **innerhalb** des konfigurierten Verzeichnisses auf.

    Ein Connector darf nur dort lesen, wofuer er eingerichtet wurde. Ohne
    diese Pruefung koennte eine Angabe wie ``../../geheim.csv`` oder ein
    absoluter Pfad aus dem Verzeichnis herausfuehren - besonders heikel,
    sobald Abfragen nicht mehr von Hand, sondern automatisiert entstehen.
    """
    if not name or not name.strip():
        raise ConnectorError("Es wurde keine Datei angegeben.")
    basis = Path(directory).resolve()
    kandidat = Path(name)
    if kandidat.is_absolute() or kandidat.drive or kandidat.anchor:
        raise ConnectorError(
            f"Nur Dateien innerhalb von {basis} sind erlaubt - "
            "ein absoluter Pfad wird nicht angenommen."
        )
    ziel = (basis / kandidat).resolve()
    try:
        ziel.relative_to(basis)
    except ValueError:
        raise ConnectorError(
            f"Nur Dateien innerhalb von {basis} sind erlaubt - "
            f"{name!r} fuehrt aus dem Verzeichnis heraus."
        ) from None
    return ziel


def sniff_dialect(sample: str) -> csv.Dialect | type[csv.Dialect]:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        return csv.excel


class CsvConnector(Connector):
    """Liest CSV-Exporte (Buchungsstapel, OP-Listen, Kontoauszuege)."""

    connector_id = "csv"
    name = "CSV-Import"
    system = "Datei"
    capabilities = ("read", "preview")
    open_questions = (
        "Welches Trennzeichen und welche Zeichenkodierung verwendet der Export?",
        "Welche Spalten enthalten Betrag, Datum, Konto, Steuerschluessel?",
        "Wird das Dezimalkomma deutsch (1.234,56) geschrieben?",
    )

    def configured(self) -> tuple[bool, str]:
        base = self.config.get("directory")
        if not base:
            return False, "Kein Verzeichnis konfiguriert (Schluessel 'directory')."
        if not Path(base).is_dir():
            return False, f"Verzeichnis existiert nicht: {base}"
        return True, f"Bereit. Verzeichnis: {base}"

    def read(self, query: str = "", limit: int = 1000, encoding: str = "", **kwargs) -> ConnectorResult:
        ok, detail = self.configured()
        if not ok:
            raise NotConfigured(f"{self.name}: {detail}")
        basis = Path(self.config["directory"])
        path = resolve_within(basis, query)
        if not path.is_file():
            available = sorted(p.name for p in basis.glob("*.csv"))
            raise ConnectorError(
                f"{self.name}: Datei '{query}' nicht gefunden. "
                f"Vorhanden: {', '.join(available) if available else 'keine CSV-Datei'}"
            )
        raw = path.read_bytes()
        for candidate in ([encoding] if encoding else []) + ["utf-8-sig", "cp1252", "latin-1"]:
            try:
                text = raw.decode(candidate)
                used = candidate
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:  # pragma: no cover - praktisch unerreichbar
            text = raw.decode("utf-8", errors="replace")
            used = "utf-8 (mit Ersatzzeichen)"

        dialect = sniff_dialect(text[:4096])
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows = []
        for index, row in enumerate(reader):
            if index >= limit:
                break
            rows.append({k: v for k, v in row.items() if k is not None})
        return ConnectorResult(
            ok=True, rows=rows,
            message=f"{len(rows)} Zeilen aus {path.name} gelesen (Kodierung {used}).",
            meta={"datei": str(path), "spalten": reader.fieldnames or [], "kodierung": used},
        )


class ExcelConnector(Connector):
    """Liest Excel-Dateien - benoetigt das optionale Paket ``openpyxl``."""

    connector_id = "excel"
    name = "Excel-Import"
    system = "Datei"
    capabilities = ("read", "preview")
    open_questions = (
        "Welches Tabellenblatt ist massgeblich?",
        "In welcher Zeile stehen die Spaltenueberschriften?",
    )

    def configured(self) -> tuple[bool, str]:
        base = self.config.get("directory")
        if not base or not Path(base).is_dir():
            return False, "Kein gueltiges Verzeichnis konfiguriert (Schluessel 'directory')."
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return False, (
                "Das optionale Paket 'openpyxl' ist nicht installiert. "
                "Excel-Dateien koennen daher nicht gelesen werden. "
                "Alternative: die Datei als CSV exportieren."
            )
        return True, f"Bereit. Verzeichnis: {base}"

    def read(self, query: str = "", sheet: str = "", header_row: int = 1,
             limit: int = 1000, **kwargs) -> ConnectorResult:
        ok, detail = self.configured()
        if not ok:
            raise NotConfigured(f"{self.name}: {detail}")
        import openpyxl  # type: ignore

        path = resolve_within(Path(self.config["directory"]), query)
        if not path.is_file():
            raise ConnectorError(f"{self.name}: Datei nicht gefunden: {path}")
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        worksheet = book[sheet] if sheet else book[book.sheetnames[0]]
        rows_iter = worksheet.iter_rows(values_only=True)
        headers: list[str] = []
        rows: list[dict] = []
        for index, values in enumerate(rows_iter, start=1):
            if index < header_row:
                continue
            if index == header_row:
                headers = [str(v) if v is not None else f"Spalte{i}" for i, v in enumerate(values)]
                continue
            if len(rows) >= limit:
                break
            rows.append(dict(zip(headers, values)))
        book.close()
        return ConnectorResult(
            ok=True, rows=rows,
            message=f"{len(rows)} Zeilen aus {path.name} / {worksheet.title} gelesen.",
            meta={"datei": str(path), "blatt": worksheet.title, "spalten": headers},
        )
